import datetime
from datetime import datetime
import os

import requests
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from PIL import Image
from django.http import JsonResponse
from django.conf import settings
from datetime import timedelta
from django.shortcuts import render, redirect
#from django.contrib.auth.models import User
from django.contrib.auth import get_user_model
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse, Http404
from django.utils.timezone import localtime
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from rest_framework import status, serializers
from rest_framework import viewsets
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import action, api_view
from rest_framework.generics import GenericAPIView
from rest_framework.viewsets import ModelViewSet, ReadOnlyModelViewSet
from rest_framework.parsers import JSONParser, FormParser
from rest_framework.response import Response
from rest_framework.serializers import Serializer
from rest_framework.views import APIView
from task.models import Task, TaskHistory
from task.serializers import TaskListSerializer, TaskSerializer, TaskHistoryListSerializer
from task.filters import CreatedBetweenFilter
from companies.models import Company
from companies.serializers import UserListSerializer
from redmail import EmailSender
from django.template import loader
from users.models import UserCustomer
# Create your views here.

User = get_user_model()

class TaskViewSet(ModelViewSet):
    #permission_classes = (IsAppAuthenticated, IsAppStaff, IsAuthenticated, IsSuperUser)
    queryset = Task.objects.all()
    serializer_class = TaskListSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_fields = ('is_enabled',)
    

    def send_email(self, request, id, *args, **kwargs):
        try:
            users = UserCustomer.objects.filter(id__in=request.data.get("user"))
            task_day = Task.objects.filter(id=id).first()

            for user in users:
                print("Usuario: ", user.email)
                if user is not None:
                    image = Image.open('C:/Users/gabri/OneDrive/Documentos/Repositorios-github/sistema_gestion_actividades/Sistema_gestion_actividades/task/templates/actividades-de-trabajo-en-equipo.png')
                    new_image = image.resize((300, 99))
                    html_msg = loader.render_to_string('C:/Users/gabri/OneDrive/Documentos/Repositorios-github/sistema_gestion_actividades/Sistema_gestion_actividades/task/templates/sendemail.html', {
                        "Usuario": " ".join(list(map(lambda x: x.capitalize(), user.username.split(" ")))),
                        "tarea": str(task_day.description),
                        'fecha_inicio': str(task_day.start_day),
                        'fecha_entrega': str(task_day.end_day),
                    })
                    html_msg = html_msg.replace('%% my_image %%', '{{ my_image }}')
                    email = EmailSender(
                        host=settings.EMAIL_HOST,
                        port=settings.EMAIL_PORT,
                        username=settings.EMAIL_HOST_USER,
                        password=settings.EMAIL_HOST_PASSWORD
                    )
                    email.send(
                        sender=settings.EMAIL_HOST_USER,
                        receivers=[user.email],
                        subject="Tarea Asignada",
                        html=html_msg,
                        body_images={"my_image": new_image}
                    )
                    print("Email enviado a: ", user.email)
            
            return Response({'access': (True), 'User': user.username}, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({'messages': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    def create(self, request, *args, **kwargs):
        serialize = TaskSerializer(data=request.data)

        if serialize.is_valid():
            start_day = serialize.validated_data['start_day']
            end_day = serialize.validated_data['end_day']
            
            start_day_actual = datetime.date(datetime.now())
            print(start_day_actual)
            #replace_data = start_day_actual.replace(microsecond=0)
            if start_day <= start_day_actual:
                return Response({'Messages': 'No puedes crear una tarea con una fecha del dia de ayer'}, status=status.HTTP_400_BAD_REQUEST)
            
            elif end_day <= start_day:
                return Response({'Messages': 'La fecha de entrega tiene que ser mayor o igual al dia que se asigno la fecha'}, status=status.HTTP_400_BAD_REQUEST)
            
            
            serialize.save()
            task_t = TaskListSerializer(instance=serialize.instance).data
            self.send_email(request, task_t["id"], task_t["user"])
            return Response(TaskListSerializer(instance=serialize.instance).data, status=status.HTTP_201_CREATED)
            #return self.send_email(request, task_t["id"], task_t["user"])
        else:
            return Response(serialize.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None):
        try:
            task_update = self.get_object()
            serializer = TaskListSerializer(task_update, data=request.data, partial=True)

            if serializer.is_valid(raise_exception=True):
                task_finish = serializer.validated_data.get('is_finished')
                if task_finish:
                    serializer.save()

                    user_ids = request.data.get('user', [])
                    users = User.objects.filter(is_active=True, id__in=user_ids)

                    for user in users:
                        history = TaskHistory(
                            task=task_update,
                            name=task_update.name,
                            description=task_update.description,
                            is_finished=task_update.is_finished,
                            departament=task_update.departament,
                            start_day=task_update.start_day,
                            end_day=task_update.end_day,
                        )
                        history.save()
                        history.user.set([user.id])

                    return Response(TaskListSerializer(instance=serializer.instance).data, status=status.HTTP_200_OK)
                else:
                    serializer.save()
                    return Response(TaskListSerializer(instance=serializer.instance).data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'messages': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class ReportTaskFinished(APIView):
    queryset = TaskHistory.objects.all()
    serializer_class = TaskHistoryListSerializer
    filter_backends = (DjangoFilterBackend, CreatedBetweenFilter)
    filter_fields = ('is_enabled',)
    
    def get(self, request, *args, **kwargs):
        # Obtener los objetos de TaskHistory según los filtros
        queryset = TaskHistory.objects.filter(is_enabled=True).order_by('-created')
        
        # Crear un nuevo archivo de Excel
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # Crear los encabezados de las columnas
        headers = ['ID', 'name', 'description', 'is_enabled', 'is_finished', 'is_pending', 'is_started', 'user', 'departament', 'start_day', 'end_day', 'created']  # Reemplaza los campos con los correctos
        for col_num, header_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet.column_dimensions[col_letter].width = 15
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = Font(bold=True)
        
        # Llenar la información de las filas
        for row_num, obj in enumerate(queryset, 2):
            sheet.cell(row=row_num, column=1).value = obj.id
            sheet.cell(row=row_num, column=3).value = obj.name
            sheet.cell(row=row_num, column=4).value = obj.description
            sheet.cell(row=row_num, column=5).value = obj.is_enabled
            sheet.cell(row=row_num, column=6).value = obj.is_pending
            sheet.cell(row=row_num, column=7).value = obj.is_started
            sheet.cell(row=row_num, column=8).value = ', '.join(obj.user.values_list('username', flat=True))  # Obtener una lista de nombres de usuarios separados por comas
            # Opción alternativa: sheet.cell(row=row_num, column=8).value = obj.user.first().username  # Obtener el primer nombre de usuario
            sheet.cell(row=row_num, column=9).value = obj.departament.name
            sheet.cell(row=row_num, column=10).value = obj.start_day
            sheet.cell(row=row_num, column=11).value = obj.end_day
        
        # Guardar el archivo de Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="task_finished.xlsx"'
        wb.save(response)
        
        return response
    
class DashboardUserViewSet(APIView):
    
    def get_queryset(self):
        company_id = self.request.query_params.get('company')
        return Company.objects.filter(company_id=company_id)
    
    def get(self, request, pk=None, format=None):
        now = datetime.now()
        companies = Company.objects.filter(id = self.request.query_params.get('company'), is_enabled=True)
        date_list = [(now - timedelta(days=x)).strftime('%Y-%m-%d') for x in range(10)]
        date_list.reverse()
        
        for company in companies:
            data = {
                'company': company.id,
                'task_finished': [],
                'task_pending': [],
                'users': [],
                'graph': []
            }
            
            if data is not None:
                task = Task.objects.filter(is_enabled=True, departament_id=company.id, start_day=now.strftime("%Y-%m-%d")).order_by('start_day')
                for tasks in task:
                    ag_data = {
                        'id': tasks.id,
                        'nombre': tasks.name,
                        'Tareas pendientes': tasks.is_pending,
                        'Tareas finalizadas': tasks.is_finished,
                        'Dia de inicio': str(tasks.start_day),
                        'ejecutivo_username': '',
                        'ejecutivo_name': ''
                    }
                    
                    #if tasks.user_id is None:
                    data['task_pending'].append(ag_data)
                        
                return Response(data, status=status.HTTP_200_OK)